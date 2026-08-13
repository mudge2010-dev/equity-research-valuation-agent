# streamlit_app.py
# Equity Research AI Agent - classroom prototype
# Free stack: Streamlit + SEC EDGAR + pandas + yfinance + openpyxl
# Generative AI is OPTIONAL: the accounting, ratio, WACC, DCF, sensitivity, and Excel modules run without an LLM.

import io
import re
import math
import json
import textwrap
from datetime import datetime, date

import numpy as np
import pandas as pd
import requests
import streamlit as st
import yfinance as yf
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------
# App configuration
# -----------------------------
st.set_page_config(page_title="Equity Research AI Agent", page_icon="📊", layout="wide")

SEC_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
SEC_SUBMISSIONS = "https://data.sec.gov/submissions/CIK{cik10}.json"
SEC_COMPANYFACTS = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik10}.json"
SEC_ARCHIVES = "https://www.sec.gov/Archives/edgar/data/{cik}/{accession}/{primary_document}"
TREASURY_URL = (
    "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/"
    "TextView?field_tdr_date_value={year}&type=daily_treasury_yield_curve"
)
DAMODARAN_ERP_URL = "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/home.htm"

M = 1_000_000.0

METRICS = {
    "Revenue": {
        "kind": "duration",
        "tags": [
            "RevenueFromContractWithCustomerExcludingAssessedTax",
            "SalesRevenueNet",
            "Revenues",
        ],
    },
    "Operating Income": {
        "kind": "duration",
        "tags": ["OperatingIncomeLoss"],
    },
    "Net Income": {
        "kind": "duration",
        "tags": [
            "NetIncomeLoss",
            "ProfitLoss",
            "NetIncomeLossAvailableToCommonStockholdersBasic",
        ],
    },
    "Operating Cash Flow": {
        "kind": "duration",
        "tags": ["NetCashProvidedByUsedInOperatingActivities"],
    },
    "Capital Expenditures": {
        "kind": "duration",
        "tags": [
            "PaymentsToAcquirePropertyPlantAndEquipment",
            "PaymentsForAdditionsToPropertyPlantAndEquipment",
            "PaymentsToAcquireProductiveAssets",
        ],
    },
    "Depreciation & Amortization": {
        "kind": "duration",
        "tags": [
            "DepreciationDepletionAndAmortization",
            "DepreciationDepletionAndAmortizationPropertyPlantAndEquipment",
            "Depreciation",
        ],
    },
    "Total Assets": {
        "kind": "instant",
        "tags": ["Assets"],
    },
    "Current Assets": {
        "kind": "instant",
        "tags": ["AssetsCurrent"],
    },
    "Current Liabilities": {
        "kind": "instant",
        "tags": ["LiabilitiesCurrent"],
    },
    "Total Liabilities": {
        "kind": "instant",
        "tags": ["Liabilities"],
    },
    "Stockholders' Equity": {
        "kind": "instant",
        "tags": [
            "StockholdersEquity",
            "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
        ],
    },
    "Long-Term Debt": {
        "kind": "instant",
        "tags": [
            "LongTermDebtNoncurrent",
            "LongTermDebt",
            "LongTermDebtAndFinanceLeaseObligationsNoncurrent",
        ],
    },
    "Inventory": {
        "kind": "instant",
        "tags": ["InventoryNet"],
    },
    "Accounts Receivable": {
        "kind": "instant",
        "tags": [
            "AccountsReceivableNetCurrent",
            "AccountsNotesAndLoansReceivableNetCurrent",
        ],
    },
    "Cash & Equivalents": {
        "kind": "instant",
        "tags": [
            "CashAndCashEquivalentsAtCarryingValue",
            "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents",
        ],
    },
    "Interest Expense": {
        "kind": "duration",
        "tags": [
            "InterestExpenseNonOperating",
            "InterestAndDebtExpense",
            "InterestExpense",
        ],
    },
    "Diluted Shares": {
        "kind": "duration_shares",
        "tags": ["WeightedAverageNumberOfDilutedSharesOutstanding"],
    },
}


# -----------------------------
# Helpers: HTTP / SEC
# -----------------------------
def sec_headers(email: str):
    clean = (email or "student@example.edu").strip()
    return {
        "User-Agent": f"EquityResearchClassroomApp/1.0 {clean}",
        "Accept-Encoding": "gzip, deflate",
        "Host": "data.sec.gov",
    }


def sec_www_headers(email: str):
    clean = (email or "student@example.edu").strip()
    return {
        "User-Agent": f"EquityResearchClassroomApp/1.0 {clean}",
        "Accept-Encoding": "gzip, deflate",
    }


@st.cache_data(ttl=24 * 3600, show_spinner=False)
def get_sec_ticker_map(email: str):
    r = requests.get(SEC_TICKERS_URL, headers=sec_www_headers(email), timeout=30)
    r.raise_for_status()
    raw = r.json()
    rows = []
    for _, x in raw.items():
        rows.append(
            {
                "ticker": str(x["ticker"]).upper(),
                "cik": int(x["cik_str"]),
                "title": x["title"],
            }
        )
    return pd.DataFrame(rows)


@st.cache_data(ttl=3600, show_spinner=False)
def get_submissions(cik: int, email: str):
    url = SEC_SUBMISSIONS.format(cik10=str(cik).zfill(10))
    r = requests.get(url, headers=sec_headers(email), timeout=30)
    r.raise_for_status()
    return r.json(), url


@st.cache_data(ttl=3600, show_spinner=False)
def get_companyfacts(cik: int, email: str):
    url = SEC_COMPANYFACTS.format(cik10=str(cik).zfill(10))
    r = requests.get(url, headers=sec_headers(email), timeout=30)
    r.raise_for_status()
    return r.json(), url


def latest_10k_from_submissions(submissions: dict, cik: int):
    recent = submissions.get("filings", {}).get("recent", {})
    df = pd.DataFrame(recent)
    if df.empty or "form" not in df:
        return None
    tenk = df[df["form"].isin(["10-K", "10-K/A"])].copy()
    if tenk.empty:
        return None
    tenk["filingDate"] = pd.to_datetime(tenk["filingDate"], errors="coerce")
    tenk = tenk.sort_values("filingDate", ascending=False)
    row = tenk.iloc[0].to_dict()
    accession_nodash = str(row["accessionNumber"]).replace("-", "")
    url = SEC_ARCHIVES.format(
        cik=cik,
        accession=accession_nodash,
        primary_document=row["primaryDocument"],
    )
    row["filing_url"] = url
    return row


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_filing_text(url: str, email: str):
    r = requests.get(url, headers=sec_www_headers(email), timeout=45)
    r.raise_for_status()
    soup = BeautifulSoup(r.content, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text)
    return text


# -----------------------------
# Helpers: XBRL mapping
# -----------------------------
def _candidate_df(companyfacts: dict, tag: str, kind: str):
    gaap = companyfacts.get("facts", {}).get("us-gaap", {})
    item = gaap.get(tag)
    if not item:
        return pd.DataFrame()
    units = item.get("units", {})
    preferred_units = ["USD"] if kind != "duration_shares" else ["shares"]
    rows = []
    for unit in preferred_units:
        rows.extend(units.get(unit, []))
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    if "form" in df:
        df = df[df["form"].isin(["10-K", "10-K/A"])]
    if "end" in df:
        df["end"] = pd.to_datetime(df["end"], errors="coerce")
    if "start" in df:
        df["start"] = pd.to_datetime(df["start"], errors="coerce")
    if "filed" in df:
        df["filed"] = pd.to_datetime(df["filed"], errors="coerce")
    if kind.startswith("duration") and "start" in df and "end" in df:
        days = (df["end"] - df["start"]).dt.days
        df = df[(days >= 300) & (days <= 430)]
    return df


def choose_metric_tag(companyfacts: dict, metric: str):
    spec = METRICS[metric]
    for tag in spec["tags"]:
        df = _candidate_df(companyfacts, tag, spec["kind"])
        if not df.empty and "val" in df:
            return tag, df
    return None, pd.DataFrame()


def build_financials(companyfacts: dict, years: int = 5):
    # Use Revenue to define the five annual fiscal endpoints when possible.
    selected = {}
    raw_dfs = {}
    for metric in METRICS:
        tag, df = choose_metric_tag(companyfacts, metric)
        selected[metric] = tag
        raw_dfs[metric] = df

    ref_metric = "Revenue" if selected.get("Revenue") else next(
        (m for m, t in selected.items() if t and METRICS[m]["kind"].startswith("duration")),
        None,
    )
    if ref_metric is None:
        return pd.DataFrame(), pd.DataFrame()

    ref = raw_dfs[ref_metric].copy()
    ref = ref.dropna(subset=["end"])
    # For duplicate facts at one period end, prefer latest filing.
    sort_cols = ["end"] + (["filed"] if "filed" in ref else [])
    ref = ref.sort_values(sort_cols)
    ref = ref.drop_duplicates(subset=["end"], keep="last")
    ends = list(ref["end"].sort_values().tail(years))
    if not ends:
        return pd.DataFrame(), pd.DataFrame()

    out = pd.DataFrame(index=[x.year for x in ends])
    out.index.name = "Fiscal Year"
    mapping_rows = []

    for metric, tag in selected.items():
        spec = METRICS[metric]
        mapping_rows.append(
            {
                "Metric": metric,
                "Selected XBRL Tag": tag or "N/A",
                "Fact Type": "Duration" if spec["kind"].startswith("duration") else "Instant",
                "Unit": "shares" if spec["kind"] == "duration_shares" else "USD",
            }
        )
        vals = []
        df = raw_dfs[metric]
        for end in ends:
            if df.empty or "end" not in df:
                vals.append(np.nan)
                continue
            candidates = df[df["end"] == end].copy()
            if candidates.empty:
                # Allow a small endpoint tolerance for unusual fiscal calendars.
                delta = (df["end"] - end).abs()
                candidates = df[delta <= pd.Timedelta(days=7)].copy()
            if candidates.empty:
                vals.append(np.nan)
                continue
            if "filed" in candidates:
                candidates = candidates.sort_values("filed")
            vals.append(pd.to_numeric(candidates.iloc[-1].get("val"), errors="coerce"))
        out[metric] = vals

    # Convert all USD metrics to millions, but diluted shares to millions of shares.
    for col in out.columns:
        out[col] = pd.to_numeric(out[col], errors="coerce") / M

    return out, pd.DataFrame(mapping_rows)


# -----------------------------
# Helpers: ratios
# -----------------------------
def safe_div(a, b):
    return np.where((pd.notna(b)) & (b != 0), a / b, np.nan)


def build_ratios(fin: pd.DataFrame):
    if fin.empty:
        return pd.DataFrame()
    r = pd.DataFrame(index=fin.index)
    r["Revenue Growth"] = fin["Revenue"].pct_change()
    r["Operating Margin"] = safe_div(fin["Operating Income"], fin["Revenue"])
    r["Net Margin"] = safe_div(fin["Net Income"], fin["Revenue"])
    r["CFO Margin"] = safe_div(fin["Operating Cash Flow"], fin["Revenue"])
    r["Capex / Revenue"] = safe_div(fin["Capital Expenditures"], fin["Revenue"])
    approx_fcf = fin["Operating Cash Flow"] - fin["Capital Expenditures"]
    r["Approx. FCF Margin"] = safe_div(approx_fcf, fin["Revenue"])
    r["Current Ratio"] = safe_div(fin["Current Assets"], fin["Current Liabilities"])
    r["Debt / Equity"] = safe_div(fin["Long-Term Debt"], fin["Stockholders' Equity"])
    r["Debt / Assets"] = safe_div(fin["Long-Term Debt"], fin["Total Assets"])

    avg_assets = (fin["Total Assets"] + fin["Total Assets"].shift(1)) / 2
    avg_equity = (fin["Stockholders' Equity"] + fin["Stockholders' Equity"].shift(1)) / 2
    r["Asset Turnover"] = safe_div(fin["Revenue"], avg_assets)
    r["ROA"] = safe_div(fin["Net Income"], avg_assets)
    r["ROE"] = safe_div(fin["Net Income"], avg_equity)
    return r


def fmt_ratio_df(r):
    out = r.copy()
    pct_cols = [
        "Revenue Growth", "Operating Margin", "Net Margin", "CFO Margin",
        "Capex / Revenue", "Approx. FCF Margin", "ROA", "ROE"
    ]
    for c in out.columns:
        if c in pct_cols:
            out[c] = out[c].map(lambda x: f"{x:.1%}" if pd.notna(x) else "N/A")
        else:
            out[c] = out[c].map(lambda x: f"{x:.2f}x" if pd.notna(x) else "N/A")
    return out


def ratio_trend_text(ratios: pd.DataFrame):
    if ratios.empty or len(ratios) < 2:
        return ["Insufficient ratio history."]
    latest = ratios.iloc[-1]
    prev = ratios.iloc[-2]
    lines = []
    for c in ratios.columns:
        if pd.isna(latest[c]) or pd.isna(prev[c]):
            continue
        delta = latest[c] - prev[c]
        direction = "increased" if delta > 0 else "declined" if delta < 0 else "was unchanged"
        if c in ["Current Ratio", "Debt / Equity", "Debt / Assets", "Asset Turnover"]:
            lines.append(f"{c} {direction} from {prev[c]:.2f}x to {latest[c]:.2f}x. Cause not established from ratio data alone.")
        else:
            lines.append(f"{c} {direction} from {prev[c]:.1%} to {latest[c]:.1%}. Cause not established from ratio data alone.")
    return lines


# -----------------------------
# Helpers: filing sections and evidence
# -----------------------------
def extract_item_section(text: str, start_item: str, end_item: str):
    if not text:
        return ""
    # Find all occurrences and choose the longest plausible segment, which tends to avoid TOC references.
    start_patterns = [
        rf"\bITEM\s+{re.escape(start_item)}[\.\:\s]",
        rf"\bItem\s+{re.escape(start_item)}[\.\:\s]",
    ]
    end_patterns = [
        rf"\bITEM\s+{re.escape(end_item)}[\.\:\s]",
        rf"\bItem\s+{re.escape(end_item)}[\.\:\s]",
    ]
    starts = []
    ends = []
    for p in start_patterns:
        starts += [m.start() for m in re.finditer(p, text)]
    for p in end_patterns:
        ends += [m.start() for m in re.finditer(p, text)]
    candidates = []
    for s in starts:
        later = [e for e in ends if e > s]
        if not later:
            continue
        e = later[0]
        seg = text[s:e]
        if 500 < len(seg) < 500_000:
            candidates.append(seg)
    return max(candidates, key=len) if candidates else ""


def evidence_snippets(text: str, keywords, max_snippets=16):
    if not text:
        return []
    sentences = re.split(r"(?<=[\.\?\!])\s+", text)
    hits = []
    seen = set()
    for i, sent in enumerate(sentences):
        s_lower = sent.lower()
        if any(k.lower() in s_lower for k in keywords):
            context = " ".join(sentences[max(0, i-1): min(len(sentences), i+2)])
            context = re.sub(r"\s+", " ", context).strip()
            if 80 <= len(context) <= 1200 and context not in seen:
                hits.append(context)
                seen.add(context)
        if len(hits) >= max_snippets:
            break
    return hits


# -----------------------------
# Helpers: market data / WACC
# -----------------------------
@st.cache_data(ttl=1800, show_spinner=False)
def get_market_data(ticker: str):
    t = yf.Ticker(ticker)
    info = {}
    try:
        info = t.get_info()
    except Exception:
        pass
    price = np.nan
    try:
        h = t.history(period="5d", auto_adjust=False)
        if not h.empty:
            price = float(h["Close"].dropna().iloc[-1])
    except Exception:
        pass
    return {
        "price": price,
        "name": info.get("longName") or info.get("shortName"),
        "sector": info.get("sector"),
        "industry": info.get("industry"),
        "marketCap": info.get("marketCap"),
        "beta_yahoo": info.get("beta"),
        "sharesOutstanding": info.get("sharesOutstanding"),
        "currency": info.get("currency"),
        "exchange": info.get("exchange"),
        "website": info.get("website"),
    }


@st.cache_data(ttl=24 * 3600, show_spinner=False)
def regression_beta(ticker: str):
    try:
        px = yf.download(
            [ticker, "^GSPC"], period="5y", interval="1mo",
            auto_adjust=True, progress=False, group_by="column", threads=False
        )
        if isinstance(px.columns, pd.MultiIndex):
            close = px["Close"]
        else:
            return np.nan, 0
        close = close.dropna(how="all")
        ret = close.pct_change().dropna()
        if ticker not in ret.columns or "^GSPC" not in ret.columns or len(ret) < 24:
            return np.nan, len(ret)
        b = ret[ticker].cov(ret["^GSPC"]) / ret["^GSPC"].var()
        return float(b), int(len(ret))
    except Exception:
        return np.nan, 0


@st.cache_data(ttl=6 * 3600, show_spinner=False)
def current_treasury_10y():
    year = date.today().year
    url = TREASURY_URL.format(year=year)
    try:
        tables = pd.read_html(url)
        candidates = [t for t in tables if "Date" in t.columns and "10 Yr" in t.columns]
        if candidates:
            t = candidates[0].copy()
            t["Date"] = pd.to_datetime(t["Date"], errors="coerce")
            t["10 Yr"] = pd.to_numeric(t["10 Yr"], errors="coerce")
            t = t.dropna(subset=["Date", "10 Yr"]).sort_values("Date")
            if not t.empty:
                row = t.iloc[-1]
                return float(row["10 Yr"]) / 100.0, row["Date"].date().isoformat(), url, "U.S. Treasury"
    except Exception:
        pass
    # Free fallback only; visibly disclosed.
    try:
        h = yf.Ticker("^TNX").history(period="5d", auto_adjust=False)
        if not h.empty:
            val = float(h["Close"].dropna().iloc[-1]) / 100.0
            return val, h.index[-1].date().isoformat(), "https://finance.yahoo.com/quote/%5ETNX/", "Yahoo ^TNX fallback"
    except Exception:
        pass
    return 0.045, "fallback", url, "Manual fallback - verify before use"


@st.cache_data(ttl=24 * 3600, show_spinner=False)
def current_damodaran_erp():
    try:
        r = requests.get(DAMODARAN_ERP_URL, timeout=30)
        r.raise_for_status()
        text = BeautifulSoup(r.content, "html.parser").get_text(" ", strip=True)
        m = re.search(r"Implied ERP on .*?=\s*([0-9]+\.[0-9]+)%", text, flags=re.I)
        if m:
            return float(m.group(1)) / 100.0, DAMODARAN_ERP_URL, "Damodaran implied ERP (first current trailing-12-month figure)"
    except Exception:
        pass
    return 0.045, DAMODARAN_ERP_URL, "Manual fallback - verify current Damodaran ERP"


# -----------------------------
# Helpers: DCF
# -----------------------------
def latest_value(fin: pd.DataFrame, col: str, default=np.nan):
    try:
        x = pd.to_numeric(fin[col], errors="coerce").dropna()
        return float(x.iloc[-1]) if len(x) else default
    except Exception:
        return default


def build_dcf(
    base_revenue,
    growths,
    ebit_margin,
    tax_rate,
    da_rev,
    capex_rev,
    nwc_rev,
    wacc,
    terminal_growth,
    exit_multiple,
    net_debt,
    diluted_shares,
):
    years = list(range(1, 6))
    rev = []
    last = base_revenue
    for g in growths:
        last = last * (1 + g)
        rev.append(last)
    d = pd.DataFrame(index=years)
    d.index.name = "Forecast Year"
    d["Revenue"] = rev
    d["Revenue Growth"] = growths
    d["EBIT Margin"] = ebit_margin
    d["EBIT"] = d["Revenue"] * ebit_margin
    d["Tax Rate"] = tax_rate
    d["EBIT(1-T)"] = d["EBIT"] * (1 - tax_rate)
    d["D&A"] = d["Revenue"] * da_rev
    d["Capex"] = d["Revenue"] * capex_rev
    d["Change in NWC"] = d["Revenue"] * nwc_rev
    d["UFCF"] = d["EBIT(1-T)"] + d["D&A"] - d["Capex"] - d["Change in NWC"]
    d["Discount Factor"] = [1 / ((1 + wacc) ** t) for t in years]
    d["PV UFCF"] = d["UFCF"] * d["Discount Factor"]
    d["EBITDA"] = d["EBIT"] + d["D&A"]

    pv_forecast = float(d["PV UFCF"].sum())
    if wacc <= terminal_growth:
        raise ValueError("WACC must be greater than terminal growth.")

    tv_pg = float(d["UFCF"].iloc[-1] * (1 + terminal_growth) / (wacc - terminal_growth))
    pv_tv_pg = tv_pg * float(d["Discount Factor"].iloc[-1])
    ev_pg = pv_forecast + pv_tv_pg
    eq_pg = ev_pg - net_debt
    px_pg = eq_pg / diluted_shares if diluted_shares > 0 else np.nan

    tv_exit = float(d["EBITDA"].iloc[-1] * exit_multiple)
    pv_tv_exit = tv_exit * float(d["Discount Factor"].iloc[-1])
    ev_exit = pv_forecast + pv_tv_exit
    eq_exit = ev_exit - net_debt
    px_exit = eq_exit / diluted_shares if diluted_shares > 0 else np.nan

    return d, {
        "PV Forecast UFCF": pv_forecast,
        "PG Terminal Value": tv_pg,
        "PG PV Terminal Value": pv_tv_pg,
        "PG Enterprise Value": ev_pg,
        "PG Equity Value": eq_pg,
        "PG Implied Share Price": px_pg,
        "PG Terminal Value / EV": pv_tv_pg / ev_pg if ev_pg else np.nan,
        "Exit Terminal Value": tv_exit,
        "Exit PV Terminal Value": pv_tv_exit,
        "Exit Enterprise Value": ev_exit,
        "Exit Equity Value": eq_exit,
        "Exit Implied Share Price": px_exit,
        "Exit Terminal Value / EV": pv_tv_exit / ev_exit if ev_exit else np.nan,
    }


def sensitivity_pg(base_revenue, growths, ebit_margin, tax_rate, da_rev, capex_rev,
                   nwc_rev, net_debt, shares, wacc_values, g_values):
    table = pd.DataFrame(index=[f"{w:.1%}" for w in wacc_values],
                         columns=[f"{g:.1%}" for g in g_values], dtype=float)
    for w in wacc_values:
        for g in g_values:
            if w <= g:
                val = np.nan
            else:
                _, out = build_dcf(base_revenue, growths, ebit_margin, tax_rate,
                                   da_rev, capex_rev, nwc_rev, w, g, 7.0, net_debt, shares)
                val = out["PG Implied Share Price"]
            table.loc[f"{w:.1%}", f"{g:.1%}"] = val
    return table


def sensitivity_exit(base_revenue, growths, ebit_margin, tax_rate, da_rev, capex_rev,
                     nwc_rev, net_debt, shares, wacc_values, multiples, terminal_growth):
    table = pd.DataFrame(index=[f"{w:.1%}" for w in wacc_values],
                         columns=[f"{m:.1f}x" for m in multiples], dtype=float)
    for w in wacc_values:
        for m in multiples:
            if w <= terminal_growth:
                val = np.nan
            else:
                _, out = build_dcf(base_revenue, growths, ebit_margin, tax_rate,
                                   da_rev, capex_rev, nwc_rev, w, terminal_growth, m,
                                   net_debt, shares)
                val = out["Exit Implied Share Price"]
            table.loc[f"{w:.1%}", f"{m:.1f}x"] = val
    return table


# -----------------------------
# Helpers: Excel model
# -----------------------------
def build_excel(
    ticker, company_name, fin, ratios, mapping, dcf_inputs, wacc_inputs,
    source_rows, current_price
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    formula_fill = PatternFill("solid", fgColor="E2F0D9")

    ws["A1"] = f"{ticker} - {company_name} | Historical Financials (USD mm; shares mm)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Metric"
    for j, yr in enumerate(fin.index, start=2):
        ws.cell(3, j, int(yr))
    for c in ws[3]:
        c.font = Font(bold=True)
        c.fill = header_fill

    for i, metric in enumerate(fin.columns, start=4):
        ws.cell(i, 1, metric)
        for j, yr in enumerate(fin.index, start=2):
            val = fin.loc[yr, metric]
            ws.cell(i, j, None if pd.isna(val) else float(val))

    # Ratio sheet with formulas linked to Historical.
    rs = wb.create_sheet("Ratios")
    rs["A1"] = "Ratio Analysis - formulas linked to Historical"
    rs["A1"].font = Font(bold=True, size=14)
    rs["A3"] = "Ratio"
    for j, yr in enumerate(fin.index, start=2):
        rs.cell(3, j, int(yr))
    for c in rs[3]:
        c.font = Font(bold=True)
        c.fill = header_fill

    hist_row = {metric: 4 + i for i, metric in enumerate(fin.columns)}
    ratio_names = list(ratios.columns)
    for i, ratio in enumerate(ratio_names, start=4):
        rs.cell(i, 1, ratio)
        for j in range(2, 2 + len(fin.index)):
            col = get_column_letter(j)
            prev_col = get_column_letter(j - 1)
            H = lambda metric, c=col: f"'Historical'!{c}{hist_row[metric]}"
            if ratio == "Revenue Growth":
                formula = None if j == 2 else f"={H('Revenue')}/'Historical'!{prev_col}{hist_row['Revenue']}-1"
            elif ratio == "Operating Margin":
                formula = f"={H('Operating Income')}/{H('Revenue')}"
            elif ratio == "Net Margin":
                formula = f"={H('Net Income')}/{H('Revenue')}"
            elif ratio == "CFO Margin":
                formula = f"={H('Operating Cash Flow')}/{H('Revenue')}"
            elif ratio == "Capex / Revenue":
                formula = f"={H('Capital Expenditures')}/{H('Revenue')}"
            elif ratio == "Approx. FCF Margin":
                formula = f"=({H('Operating Cash Flow')}-{H('Capital Expenditures')})/{H('Revenue')}"
            elif ratio == "Current Ratio":
                formula = f"={H('Current Assets')}/{H('Current Liabilities')}"
            elif ratio == "Debt / Equity":
                equity_key = "Stockholders\' Equity"
                formula = f"={H('Long-Term Debt')}/{H(equity_key)}"
            elif ratio == "Debt / Assets":
                formula = f"={H('Long-Term Debt')}/{H('Total Assets')}"
            elif ratio == "Asset Turnover":
                formula = None if j == 2 else f"={H('Revenue')}/AVERAGE({H('Total Assets')},'Historical'!{prev_col}{hist_row['Total Assets']})"
            elif ratio == "ROA":
                formula = None if j == 2 else f"={H('Net Income')}/AVERAGE({H('Total Assets')},'Historical'!{prev_col}{hist_row['Total Assets']})"
            elif ratio == "ROE":
                equity_key = "Stockholders\' Equity"
                formula = None if j == 2 else f"={H('Net Income')}/AVERAGE({H(equity_key)},\'Historical\'!{prev_col}{hist_row[equity_key]})"
            else:
                formula = None
            if formula:
                rs.cell(i, j, formula)
                rs.cell(i, j).fill = formula_fill

    # WACC sheet.
    ww = wb.create_sheet("WACC")
    ww["A1"] = "WACC Calculation"
    ww["A1"].font = Font(bold=True, size=14)
    wacc_rows = [
        ("Risk-free rate", wacc_inputs["rf"]),
        ("Beta", wacc_inputs["beta"]),
        ("Equity risk premium", wacc_inputs["erp"]),
        ("Pre-tax cost of debt", wacc_inputs["cost_debt"]),
        ("Tax rate", wacc_inputs["tax"]),
        ("Market value of equity ($mm)", wacc_inputs["market_equity"]),
        ("Debt ($mm)", wacc_inputs["debt"]),
    ]
    for i, (label, value) in enumerate(wacc_rows, start=3):
        ww.cell(i, 1, label)
        ww.cell(i, 2, value)
        ww.cell(i, 2).fill = input_fill
    ww["A11"] = "Cost of Equity (CAPM)"
    ww["B11"] = "=B3+B4*B5"
    ww["A12"] = "Equity Weight"
    ww["B12"] = "=B8/(B8+B9)"
    ww["A13"] = "Debt Weight"
    ww["B13"] = "=B9/(B8+B9)"
    ww["A14"] = "After-tax Cost of Debt"
    ww["B14"] = "=B6*(1-B7)"
    ww["A15"] = "WACC"
    ww["B15"] = "=B12*B11+B13*B14"
    for cell in ["B11", "B12", "B13", "B14", "B15"]:
        ww[cell].fill = formula_fill

    # DCF.
    ds = wb.create_sheet("DCF")
    ds["A1"] = "Five-Year UFCF DCF - editable yellow cells; formulas in green"
    ds["A1"].font = Font(bold=True, size=14)
    assumptions = [
        ("Base Revenue ($mm)", dcf_inputs["base_revenue"]),
        ("Year 1 Revenue Growth", dcf_inputs["growths"][0]),
        ("Year 2 Revenue Growth", dcf_inputs["growths"][1]),
        ("Year 3 Revenue Growth", dcf_inputs["growths"][2]),
        ("Year 4 Revenue Growth", dcf_inputs["growths"][3]),
        ("Year 5 Revenue Growth", dcf_inputs["growths"][4]),
        ("EBIT Margin", dcf_inputs["ebit_margin"]),
        ("Tax Rate", dcf_inputs["tax_rate"]),
        ("D&A / Revenue", dcf_inputs["da_rev"]),
        ("Capex / Revenue", dcf_inputs["capex_rev"]),
        ("Change in NWC / Revenue", dcf_inputs["nwc_rev"]),
        ("Terminal Growth", dcf_inputs["terminal_growth"]),
        ("Exit EV/EBITDA Multiple", dcf_inputs["exit_multiple"]),
        ("Net Debt ($mm)", dcf_inputs["net_debt"]),
        ("Diluted Shares (mm)", dcf_inputs["shares"]),
        ("WACC (linked from WACC)", None),
    ]
    for i, (label, value) in enumerate(assumptions, start=3):
        ds.cell(i, 1, label)
        if i == 18:
            ds.cell(i, 2, "=WACC!B15")
            ds.cell(i, 2).fill = formula_fill
        else:
            ds.cell(i, 2, value)
            ds.cell(i, 2).fill = input_fill

    # Forecast columns B:F
    start_row = 21
    ds.cell(start_row, 1, "Forecast")
    for j, yr in enumerate(range(1, 6), start=2):
        ds.cell(start_row, j, f"Year {yr}")
        ds.cell(start_row, j).fill = header_fill

    row_map = {
        "Revenue Growth": 22,
        "Revenue": 23,
        "EBIT Margin": 24,
        "EBIT": 25,
        "Tax Rate": 26,
        "EBIT(1-T)": 27,
        "D&A / Revenue": 28,
        "D&A": 29,
        "Capex / Revenue": 30,
        "Capex": 31,
        "Change in NWC / Revenue": 32,
        "Change in NWC": 33,
        "UFCF": 34,
        "Discount Factor": 35,
        "PV UFCF": 36,
        "EBITDA": 37,
    }
    for name, rr in row_map.items():
        ds.cell(rr, 1, name)

    for j in range(2, 7):
        c = get_column_letter(j)
        prev = get_column_letter(j - 1)
        g_assumption_row = 3 + (j - 1)  # B4:B8 for years 1-5
        ds.cell(22, j, f"=$B${g_assumption_row}")
        if j == 2:
            ds.cell(23, j, f"=$B$3*(1+{c}22)")
        else:
            ds.cell(23, j, f"={prev}23*(1+{c}22)")
        ds.cell(24, j, "=$B$9")
        ds.cell(25, j, f"={c}23*{c}24")
        ds.cell(26, j, "=$B$10")
        ds.cell(27, j, f"={c}25*(1-{c}26)")
        ds.cell(28, j, "=$B$11")
        ds.cell(29, j, f"={c}23*{c}28")
        ds.cell(30, j, "=$B$12")
        ds.cell(31, j, f"={c}23*{c}30")
        ds.cell(32, j, "=$B$13")
        ds.cell(33, j, f"={c}23*{c}32")
        ds.cell(34, j, f"={c}27+{c}29-{c}31-{c}33")
        ds.cell(35, j, f"=1/(1+$B$18)^{j-1}")
        ds.cell(36, j, f"={c}34*{c}35")
        ds.cell(37, j, f"={c}25+{c}29")
        for rr in range(22, 38):
            ds.cell(rr, j).fill = formula_fill

    # Valuation formulas.
    out_row = 40
    outputs = [
        ("PV of Forecast UFCF", "=SUM(B36:F36)"),
        ("PG Terminal Value", "=F34*(1+$B$14)/($B$18-$B$14)"),
        ("PG PV Terminal Value", "=B41*F35"),
        ("PG Enterprise Value", "=B40+B42"),
        ("PG Equity Value", "=B43-$B$16"),
        ("PG Implied Share Price", "=B44/$B$17"),
        ("PG PV Terminal Value / EV", "=B42/B43"),
        ("Exit Terminal Value", "=F37*$B$15"),
        ("Exit PV Terminal Value", "=B47*F35"),
        ("Exit Enterprise Value", "=B40+B48"),
        ("Exit Equity Value", "=B49-$B$16"),
        ("Exit Implied Share Price", "=B50/$B$17"),
        ("Exit PV Terminal Value / EV", "=B48/B49"),
        ("Current Market Price", current_price if pd.notna(current_price) else None),
        ("PG Upside / (Downside)", "=B45/B53-1"),
        ("Exit Upside / (Downside)", "=B51/B53-1"),
    ]
    for i, (label, formula_or_value) in enumerate(outputs, start=out_row):
        ds.cell(i, 1, label)
        ds.cell(i, 2, formula_or_value)
        if isinstance(formula_or_value, str) and formula_or_value.startswith("="):
            ds.cell(i, 2).fill = formula_fill

    # XBRL mapping.
    ms = wb.create_sheet("XBRL Mapping")
    for j, col in enumerate(mapping.columns, start=1):
        ms.cell(1, j, col).font = Font(bold=True)
        ms.cell(1, j).fill = header_fill
    for i, row in mapping.reset_index(drop=True).iterrows():
        for j, col in enumerate(mapping.columns, start=1):
            ms.cell(i + 2, j, row[col])

    # Sources.
    ss = wb.create_sheet("Sources")
    source_df = pd.DataFrame(source_rows)
    if source_df.empty:
        source_df = pd.DataFrame(columns=["Source", "URL", "Section/Page", "Evidence/Note"])
    for j, col in enumerate(source_df.columns, start=1):
        ss.cell(1, j, col).font = Font(bold=True)
        ss.cell(1, j).fill = header_fill
    for i, row in source_df.reset_index(drop=True).iterrows():
        for j, col in enumerate(source_df.columns, start=1):
            ss.cell(i + 2, j, row[col])

    # Formatting.
    for sheet in wb.worksheets:
        sheet.freeze_panes = "B4" if sheet.title in ["Historical", "Ratios"] else "A2"
        for col in range(1, min(sheet.max_column, 12) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 22 if col == 1 else 16
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
    for sheet_name in ["Ratios", "WACC", "DCF"]:
        sh = wb[sheet_name]
        for row in sh.iter_rows():
            for cell in row:
                if cell.row in range(3, 19) and isinstance(cell.value, float) and abs(cell.value) < 1:
                    cell.number_format = "0.0%"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


# -----------------------------
# Helpers: prompts / audit
# -----------------------------
def dataframe_markdown(df, decimals=3):
    if df is None or df.empty:
        return "N/A"
    return df.round(decimals).to_markdown()


def build_ratio_prompt(company_name, ticker, ratios):
    return f"""You are assisting a finance student with ratio TREND analysis for {company_name} ({ticker}).

Use ONLY the ratio table below. Identify what changed; do not invent why it changed.
Analyze profitability, liquidity, leverage, efficiency, cash flow, returns, important trends,
and questions requiring further investigation.

RULES:
- Never infer a business cause from ratios alone.
- For any causal question not established by this table, write exactly:
  "Cause not established from ratio data alone."
- Do not invent peers, industry averages, management commentary, guidance, or market data.
- Distinguish observation from interpretation.

RATIO TABLE
{dataframe_markdown(ratios)}
"""


def build_memo_prompt(company_name, ticker, company_info, fin, ratios, dcf_inputs, dcf_outputs,
                      revenue_evidence, risk_evidence, filing_url, sources):
    fin_text = dataframe_markdown(fin)
    ratio_text = dataframe_markdown(ratios)
    rev_text = "\n\n".join([f"[Revenue evidence {i+1}] {x}" for i, x in enumerate(revenue_evidence)]) or "N/A"
    risk_text = "\n\n".join([f"[Risk evidence {i+1}] {x}" for i, x in enumerate(risk_evidence)]) or "N/A"
    return f"""ROLE
You are a junior equity-research drafting assistant. Draft a FIRST-DRAFT equity research memo
for {company_name} ({ticker}). The human student is the analyst of record and must verify every
material statement.

SOURCE BOUNDARY
Use ONLY the supplied evidence and calculations below. Do not browse. Do not add outside facts.
If evidence is insufficient, write exactly: "Not established from supplied evidence."

DO NOT INVENT
Management guidance; management quotations; market share; peer multiples; industry benchmarks;
catalysts; causal explanations; forecasts not supplied by the analyst.

REQUIRED SECTIONS
1. Investment Snapshot
2. Five-Year Financial Review
3. Ratio Analysis
4. Revenue Drivers
5. Risk Factors
6. DCF Valuation and WACC Calculation
7. Sensitivity Analysis
8. Conclusion
9. Verification Items

WRITING CONTROLS
- Label analyst/model assumptions as assumptions, never as company expectations.
- Historical facts must tie to supplied financials or filing evidence.
- Calculations must tie to supplied tables.
- Causal statements require explicit filing evidence.
- If a ratio changed but cause is not evidenced, state: "Cause not established from ratio data alone."
- Use ranges and sensitivity language rather than false valuation precision.
- Every management-attribution statement must cite the supplied primary-source excerpt.
- Add [VERIFY] beside any statement that still needs manual checking.

COMPANY INFORMATION
{json.dumps(company_info, indent=2, default=str)}

FILING
{filing_url}

FIVE-YEAR FINANCIAL STATEMENTS (USD mm unless noted)
{fin_text}

RATIO ANALYSIS
{ratio_text}

DCF ASSUMPTIONS
{json.dumps(dcf_inputs, indent=2, default=str)}

DCF OUTPUTS
{json.dumps(dcf_outputs, indent=2, default=str)}

REVENUE-DRIVER SEC EVIDENCE
{rev_text}

RISK-FACTOR SEC EVIDENCE
{risk_text}

SOURCE LINKS
{json.dumps(sources, indent=2, default=str)}
"""


def claim_candidates(memo: str, limit=20):
    if not memo or not memo.strip():
        return pd.DataFrame(columns=[
            "Claim #", "AI Claim", "Claim Type", "Evidence",
            "Verification Status", "Correction"
        ])
    sentences = re.split(r"(?<=[\.\?\!])\s+", re.sub(r"\s+", " ", memo.strip()))
    trigger = re.compile(
        r"(\d|%|\$|because|due to|driven by|primarily|reflecting|resulted from|"
        r"management expects|management believes|management indicated|company expects|"
        r"guidance suggests|management anticipates|wacc|terminal|upside|downside|margin|revenue)",
        re.I,
    )
    chosen = []
    for s in sentences:
        if len(s) >= 35 and trigger.search(s):
            chosen.append(s.strip())
        if len(chosen) >= limit:
            break
    if len(chosen) < min(limit, len(sentences)):
        for s in sentences:
            s = s.strip()
            if len(s) >= 35 and s not in chosen:
                chosen.append(s)
            if len(chosen) >= limit:
                break
    return pd.DataFrame({
        "Claim #": range(1, len(chosen) + 1),
        "AI Claim": chosen,
        "Claim Type": "",
        "Evidence": "",
        "Verification Status": "",
        "Correction": "",
    })


def second_audit_prompt(memo: str):
    return f"""Perform a PRELIMINARY claim audit of the memo below. This is not final verification.

For approximately 15-20 material claims:
1. Classify Claim Type as one of:
Historical Fact; Calculation; Management Statement; Analytical Interpretation;
Forecast Assumption; Valuation Output.
2. Assign a preliminary Verification Status:
VERIFIED; PARTIALLY SUPPORTED / QUALIFY; UNSUPPORTED; CONTRADICTED.
3. Flag causal terms: because, due to, driven by, primarily, reflecting, resulted from.
4. Flag management-attribution terms: management expects/believes/indicated/anticipates,
the company expects, guidance suggests.
5. Flag every material number for manual checking.
6. For forecast language, ask "Expected by whom?" and distinguish management guidance from
the student's model assumption.
7. Never certify your own audit as final. End with:
"Student must manually verify each material claim against the primary source or model."

MEMO
{memo}
"""


# -----------------------------
# Sidebar / ticker
# -----------------------------
st.title("📊 Equity Research AI Agent")
st.caption("Classroom prototype: evidence + calculations → AI synthesis → student verification + judgment")

with st.sidebar:
    st.header("Run Settings")
    ticker = st.text_input("U.S. ticker", value="STLD").strip().upper()
    sec_email = st.text_input(
        "SEC User-Agent contact email",
        value="student@example.edu",
        help="SEC requests a descriptive User-Agent. Replace this classroom placeholder with your school email.",
    )
    st.caption("Changing the ticker reruns the workflow.")
    st.divider()
    st.markdown("**Governance rule:** Do not treat AI output as a source.")

if not ticker:
    st.error("Enter a ticker.")
    st.stop()

# Reset reconciliation gate when ticker changes.
if st.session_state.get("active_ticker") != ticker:
    st.session_state["active_ticker"] = ticker
    for key in list(st.session_state.keys()):
        if key.startswith("recon_"):
            st.session_state[key] = False

# -----------------------------
# Load company / SEC / market
# -----------------------------
load_errors = []
try:
    ticker_map = get_sec_ticker_map(sec_email)
    match = ticker_map[ticker_map["ticker"] == ticker]
    if match.empty:
        st.error(f"{ticker} was not found in the SEC ticker map.")
        st.stop()
    cik = int(match.iloc[0]["cik"])
    sec_title = str(match.iloc[0]["title"])
except Exception as e:
    st.error(f"SEC ticker lookup failed: {e}")
    st.stop()

try:
    submissions, submissions_url = get_submissions(cik, sec_email)
    latest_10k = latest_10k_from_submissions(submissions, cik)
except Exception as e:
    submissions, latest_10k, submissions_url = {}, None, ""
    load_errors.append(f"SEC submissions failed: {e}")

try:
    facts, facts_url = get_companyfacts(cik, sec_email)
    fin, mapping = build_financials(facts, years=5)
except Exception as e:
    facts, facts_url, fin, mapping = {}, "", pd.DataFrame(), pd.DataFrame()
    load_errors.append(f"SEC Companyfacts failed: {e}")

try:
    market = get_market_data(ticker)
except Exception as e:
    market = {}
    load_errors.append(f"Market data failed: {e}")

company_name = market.get("name") or sec_title
filing_url = latest_10k.get("filing_url") if latest_10k else None

filing_text = ""
if filing_url:
    try:
        filing_text = fetch_filing_text(filing_url, sec_email)
    except Exception as e:
        load_errors.append(f"10-K text retrieval failed: {e}")

ratios = build_ratios(fin) if not fin.empty else pd.DataFrame()

tabs = st.tabs([
    "1. Company",
    "2. Financial Statements",
    "3. Ratio Analysis",
    "4. Revenue Drivers",
    "5. Risks",
    "6. DCF-WACC Calculation",
    "7. Excel",
    "8. Memo",
    "9. Audit",
])


# -----------------------------
# Tab 1 Company
# -----------------------------
with tabs[0]:
    st.subheader("Module 1 — Company Agent")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Ticker", ticker)
    c2.metric("SEC CIK", str(cik).zfill(10))
    c3.metric("Company", company_name)
    px = market.get("price", np.nan)
    c4.metric("Market Price Reference", f"${px:,.2f}" if pd.notna(px) else "N/A")

    info_table = pd.DataFrame({
        "Field": ["SEC name", "Sector", "Industry", "Exchange", "Currency", "Website"],
        "Value": [
            sec_title, market.get("sector"), market.get("industry"),
            market.get("exchange"), market.get("currency"), market.get("website")
        ],
    })
    st.dataframe(info_table, use_container_width=True, hide_index=True)

    if latest_10k:
        st.success(
            f"Latest 10-K identified: filed {latest_10k.get('filingDate')} "
            f"(report date {latest_10k.get('reportDate')})."
        )
        st.link_button("Open latest SEC 10-K", filing_url)
    else:
        st.warning("Latest 10-K could not be identified automatically.")

    st.markdown("**Primary data links**")
    st.write(f"SEC submissions JSON: {submissions_url}")
    st.write(f"SEC Companyfacts JSON: {facts_url}")
    if filing_url:
        st.write(f"Latest 10-K: {filing_url}")

    if load_errors:
        with st.expander("Data retrieval warnings"):
            for e in load_errors:
                st.warning(e)

    st.info(
        "Market price and beta references are convenience data for research/education. "
        "The accounting model is built from SEC filings. Verify market data before submission."
    )


# -----------------------------
# Tab 2 Financial Statements
# -----------------------------
with tabs[1]:
    st.subheader("Module 2 — Financial Statements Agent")
    st.markdown(
        "SEC Companyfacts/XBRL is used for approximately five annual fiscal years. "
        "**Duration** facts cover a period (income statement/cash flow); **Instant** facts are point-in-time "
        "(balance sheet). Values below are USD millions; diluted shares are millions of shares."
    )

    if fin.empty:
        st.error("No usable annual Companyfacts data were mapped. Stop here and fix the accounting-data mapping.")
    else:
        display_fin = fin.copy()
        st.dataframe(display_fin.style.format("{:,.1f}", na_rep="N/A"), use_container_width=True)
        st.download_button(
            "Download financial statements CSV",
            data=fin.to_csv().encode("utf-8"),
            file_name=f"{ticker}_historical_financials.csv",
            mime="text/csv",
        )

        st.markdown("#### XBRL tag audit")
        st.dataframe(mapping, use_container_width=True, hide_index=True)
        st.warning(
            "N/A can mean the company used a different XBRL tag or disclosure structure. "
            "N/A is not automatically zero."
        )

        st.markdown("#### Required manual reconciliation gate")
        st.write("Open the latest 10-K and reconcile these four metrics before later modules are used.")
        r1 = st.checkbox("Revenue reconciled to 10-K", key=f"recon_rev_{ticker}")
        r2 = st.checkbox("Operating Income reconciled to 10-K", key=f"recon_oi_{ticker}")
        r3 = st.checkbox("Operating Cash Flow reconciled to 10-K", key=f"recon_cfo_{ticker}")
        r4 = st.checkbox("Capital Expenditures reconciled to 10-K", key=f"recon_capex_{ticker}")
        reconciled = bool(r1 and r2 and r3 and r4)
        if reconciled:
            st.success("Accounting-data reconciliation gate passed for this classroom run.")
        else:
            st.error("Gate not passed. Do not rely on Ratio, Revenue Driver, Risk, DCF, Excel, Memo, or Audit outputs yet.")

        if filing_url:
            st.link_button("Open 10-K for reconciliation", filing_url)


def gate_passed():
    return all([
        st.session_state.get(f"recon_rev_{ticker}", False),
        st.session_state.get(f"recon_oi_{ticker}", False),
        st.session_state.get(f"recon_cfo_{ticker}", False),
        st.session_state.get(f"recon_capex_{ticker}", False),
    ])


# -----------------------------
# Tab 3 Ratios
# -----------------------------
with tabs[2]:
    st.subheader("Module 3 — Ratio Analysis Agent")
    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete. Return to Financial Statements and reconcile first.")
    elif ratios.empty:
        st.error("Ratios cannot be calculated.")
    else:
        st.dataframe(fmt_ratio_df(ratios), use_container_width=True)
        st.markdown("#### Constrained trend statements")
        for line in ratio_trend_text(ratios):
            st.write("• " + line)

        st.markdown("#### Manual ratio verification")
        latest_year = ratios.index[-1]
        st.code(
            "Operating Margin = Operating Income / Revenue\n"
            "Net Margin = Net Income / Revenue\n"
            "Current Ratio = Current Assets / Current Liabilities\n"
            "Debt / Equity = Long-Term Debt / Stockholders' Equity\n"
            "Approximate FCF Margin = (Operating Cash Flow - Capex) / Revenue"
        )
        st.caption(f"Manually recalculate several ratios for fiscal year {latest_year}.")

        ratio_prompt = build_ratio_prompt(company_name, ticker, ratios)
        st.markdown("#### Optional Ratio Analysis AI prompt")
        st.text_area("Copy into an approved AI tool", ratio_prompt, height=360)


# -----------------------------
# Tab 4 Revenue Drivers
# -----------------------------
with tabs[3]:
    st.subheader("Module 4 — Revenue Driver Agent")
    revenue_evidence = []
    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete.")
    else:
        item1 = extract_item_section(filing_text, "1", "1A")
        item7 = extract_item_section(filing_text, "7", "7A")
        rev_source_text = (item1 + " " + item7).strip() or filing_text
        revenue_evidence = evidence_snippets(
            rev_source_text,
            ["price", "pricing", "volume", "shipment", "mix", "demand", "capacity",
             "sales", "revenue", "tons", "operations"],
            max_snippets=16,
        )
        st.markdown("Conceptual framework: **Revenue = Price × Volume × Mix**.")
        st.write(
            "The excerpts below are evidence candidates, not automatic causal conclusions. "
            "Students must identify the exact filing support for each material causal statement."
        )
        if revenue_evidence:
            for i, x in enumerate(revenue_evidence, start=1):
                st.markdown(f"**Evidence candidate {i}**")
                st.write(x)
        else:
            st.warning("Automatic evidence extraction was inconclusive. Use the 10-K link and paste evidence manually.")

        manual_rev = st.text_area(
            "Student revenue-driver notes / manually verified filing excerpts",
            key=f"manual_rev_{ticker}",
            height=180,
            placeholder="Observed trend | Management explanation with source | Your interpretation",
        )
        st.info(
            "Keep three layers separate: (1) observed financial trend, "
            "(2) management explanation supported by the filing, and (3) analyst interpretation."
        )


# -----------------------------
# Tab 5 Risks
# -----------------------------
with tabs[4]:
    st.subheader("Module 5 — Risk Agent")
    risk_evidence = []
    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete.")
    else:
        item1a = extract_item_section(filing_text, "1A", "1B")
        if not item1a:
            item1a = extract_item_section(filing_text, "1A", "2")
        risk_evidence = evidence_snippets(
            item1a or filing_text,
            ["risk", "commodity", "price", "demand", "cyber", "environment", "regulation",
             "competition", "supply", "energy", "labor", "capital", "debt", "interest"],
            max_snippets=12,
        )
        st.write("Use Item 1A to select approximately 4–6 material risks.")
        if risk_evidence:
            for i, x in enumerate(risk_evidence, start=1):
                st.markdown(f"**Risk evidence candidate {i}**")
                st.write(x)
        else:
            st.warning("Automatic Item 1A extraction was inconclusive; use the SEC filing directly.")

        st.markdown("#### Financial transmission mechanism")
        st.code(
            "Risk event → operating/financial variable → UFCF effect → DCF effect\n"
            "Example: Commodity price decline → lower selling prices → lower revenue → "
            "lower EBIT margin → lower UFCF → lower DCF value"
        )
        manual_risk = st.text_area(
            "Student-selected 4–6 risks and transmission mechanisms",
            key=f"manual_risk_{ticker}",
            height=180,
        )


# -----------------------------
# Tab 6 DCF / WACC
# -----------------------------
with tabs[5]:
    st.subheader("Module 6 — DCF Agent and WACC Calculation")
    dcf_ready = False
    dcf_df = pd.DataFrame()
    dcf_out = {}
    pg_sens = pd.DataFrame()
    exit_sens = pd.DataFrame()
    dcf_inputs = {}
    wacc_inputs = {}

    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete.")
    elif fin.empty:
        st.error("Historical financials unavailable.")
    else:
        rf_default, rf_date, rf_url, rf_source = current_treasury_10y()
        erp_default, erp_url, erp_source = current_damodaran_erp()
        beta_reg, beta_obs = regression_beta(ticker)
        beta_yahoo = market.get("beta_yahoo")
        beta_default = float(beta_yahoo) if beta_yahoo is not None and pd.notna(beta_yahoo) else (
            float(beta_reg) if pd.notna(beta_reg) else 1.0
        )

        revenue_latest = latest_value(fin, "Revenue", 1000.0)
        op_margin_hist = latest_value(ratios, "Operating Margin", 0.10) if not ratios.empty else 0.10
        da_rev_hist = latest_value(fin, "Depreciation & Amortization", 0.0) / revenue_latest if revenue_latest else 0.03
        capex_rev_hist = latest_value(fin, "Capital Expenditures", 0.0) / revenue_latest if revenue_latest else 0.04
        debt_latest = latest_value(fin, "Long-Term Debt", 0.0)
        cash_latest = latest_value(fin, "Cash & Equivalents", 0.0)
        net_debt_default = debt_latest - cash_latest
        shares_xbrl = latest_value(fin, "Diluted Shares", np.nan)
        shares_yahoo = (market.get("sharesOutstanding") or np.nan) / M if market.get("sharesOutstanding") else np.nan
        shares_default = shares_xbrl if pd.notna(shares_xbrl) and shares_xbrl > 0 else shares_yahoo
        if pd.isna(shares_default) or shares_default <= 0:
            shares_default = 100.0

        market_equity_default = (market.get("marketCap") or np.nan) / M if market.get("marketCap") else (
            (market.get("price") or 0) * shares_default
        )
        if not pd.notna(market_equity_default) or market_equity_default <= 0:
            market_equity_default = 10_000.0

        st.markdown("#### WACC inputs")
        st.caption(
            f"Risk-free source: {rf_source}, observation {rf_date}. "
            f"ERP source: {erp_source}. Beta cross-check: 5-year monthly {ticker} vs. S&P 500 regression "
            f"= {beta_reg:.2f} using {beta_obs} observations." if pd.notna(beta_reg) else
            f"Risk-free source: {rf_source}, observation {rf_date}. ERP source: {erp_source}."
        )
        st.write(f"Treasury source: {rf_url}")
        st.write(f"ERP source: {erp_url}")

        w1, w2, w3, w4 = st.columns(4)
        rf = w1.number_input("Risk-free rate", value=float(rf_default), step=0.001, format="%.4f")
        beta = w2.number_input("Beta", value=float(beta_default), step=0.05, format="%.2f")
        erp = w3.number_input("Equity risk premium", value=float(erp_default), step=0.001, format="%.4f")
        pretax_cod = w4.number_input("Pre-tax cost of debt", value=0.055, step=0.005, format="%.4f")

        w5, w6, w7 = st.columns(3)
        wacc_tax = w5.number_input("WACC tax rate", value=0.24, step=0.01, format="%.3f")
        market_equity = w6.number_input("Market equity value ($mm)", value=float(market_equity_default), step=100.0)
        debt_for_wacc = w7.number_input("Debt for WACC ($mm)", value=float(max(debt_latest, 0.0)), step=100.0)

        cost_equity = rf + beta * erp
        total_capital = market_equity + debt_for_wacc
        e_weight = market_equity / total_capital if total_capital > 0 else 1.0
        d_weight = debt_for_wacc / total_capital if total_capital > 0 else 0.0
        wacc_calc = e_weight * cost_equity + d_weight * pretax_cod * (1 - wacc_tax)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Cost of Equity (CAPM)", f"{cost_equity:.2%}")
        m2.metric("Equity Weight", f"{e_weight:.1%}")
        m3.metric("Debt Weight", f"{d_weight:.1%}")
        m4.metric("Calculated WACC", f"{wacc_calc:.2%}")

        st.markdown("#### Five-year UFCF assumptions")
        gcols = st.columns(5)
        growths = [
            gcols[i].number_input(
                f"Year {i+1} Revenue Growth", value=0.03, step=0.01, format="%.3f",
                key=f"g{i+1}_{ticker}"
            )
            for i in range(5)
        ]
        a1, a2, a3, a4 = st.columns(4)
        ebit_margin = a1.number_input("EBIT Margin", value=float(op_margin_hist if pd.notna(op_margin_hist) else 0.10), step=0.01, format="%.3f")
        tax_rate = a2.number_input("DCF Tax Rate", value=float(wacc_tax), step=0.01, format="%.3f")
        da_rev = a3.number_input("D&A / Revenue", value=float(da_rev_hist if pd.notna(da_rev_hist) and da_rev_hist >= 0 else 0.03), step=0.005, format="%.3f")
        capex_rev = a4.number_input("Capex / Revenue", value=float(capex_rev_hist if pd.notna(capex_rev_hist) and capex_rev_hist >= 0 else 0.04), step=0.005, format="%.3f")

        b1, b2, b3, b4 = st.columns(4)
        nwc_rev = b1.number_input("Change in NWC / Revenue", value=0.01, step=0.005, format="%.3f",
                                  help="Classroom simplification: forecast annual change in NWC as a percentage of revenue.")
        wacc = b2.number_input("DCF WACC", value=float(wacc_calc), step=0.005, format="%.4f")
        terminal_growth = b3.number_input("Terminal Growth", value=0.025, step=0.005, format="%.4f")
        exit_multiple = b4.number_input("Exit EV/EBITDA Multiple", value=7.0, step=0.5, format="%.1f")

        c1, c2, c3 = st.columns(3)
        net_debt = c1.number_input("Net Debt ($mm)", value=float(net_debt_default), step=100.0)
        diluted_shares = c2.number_input("Diluted Shares Outstanding (mm)", value=float(shares_default), step=1.0)
        base_revenue = c3.number_input("Base Revenue ($mm)", value=float(revenue_latest), step=100.0)

        st.caption(
            "Exit-multiple terminal method uses EV/EBITDA because this is an unlevered enterprise-value DCF. "
            "A P/E 'equity multiple' would be levered and is not directly consistent with UFCF."
        )

        if wacc <= terminal_growth:
            st.error("Model control failed: WACC must be greater than terminal growth.")
        else:
            try:
                dcf_df, dcf_out = build_dcf(
                    base_revenue, growths, ebit_margin, tax_rate, da_rev, capex_rev,
                    nwc_rev, wacc, terminal_growth, exit_multiple, net_debt, diluted_shares
                )
                dcf_ready = True
            except Exception as e:
                st.error(str(e))

        if dcf_ready:
            display_dcf = dcf_df.copy()
            st.dataframe(display_dcf.style.format({
                "Revenue": "{:,.1f}", "Revenue Growth": "{:.1%}", "EBIT Margin": "{:.1%}",
                "EBIT": "{:,.1f}", "Tax Rate": "{:.1%}", "EBIT(1-T)": "{:,.1f}",
                "D&A": "{:,.1f}", "Capex": "{:,.1f}", "Change in NWC": "{:,.1f}",
                "UFCF": "{:,.1f}", "Discount Factor": "{:.3f}", "PV UFCF": "{:,.1f}",
                "EBITDA": "{:,.1f}",
            }), use_container_width=True)

            o1, o2, o3, o4 = st.columns(4)
            o1.metric("PG Implied Price", f"${dcf_out['PG Implied Share Price']:,.2f}")
            o2.metric("Exit-Multiple Implied Price", f"${dcf_out['Exit Implied Share Price']:,.2f}")
            o3.metric("PG Terminal PV / EV", f"{dcf_out['PG Terminal Value / EV']:.1%}")
            o4.metric("Exit Terminal PV / EV", f"{dcf_out['Exit Terminal Value / EV']:.1%}")

            if dcf_out["PG Terminal Value / EV"] > 0.75 or dcf_out["Exit Terminal Value / EV"] > 0.75:
                st.warning(
                    "Terminal value exceeds 75% of enterprise value under at least one method. "
                    "Treat the valuation as highly assumption-sensitive and explain this limitation."
                )

            st.markdown("#### Sensitivity analysis — WACC / terminal growth")
            wacc_values = np.linspace(max(0.01, wacc - 0.02), wacc + 0.02, 5)
            g_values = np.linspace(max(0.0, terminal_growth - 0.01), terminal_growth + 0.01, 5)
            pg_sens = sensitivity_pg(
                base_revenue, growths, ebit_margin, tax_rate, da_rev, capex_rev,
                nwc_rev, net_debt, diluted_shares, wacc_values, g_values
            )
            st.dataframe(pg_sens.style.format("${:,.2f}", na_rep="N/M"), use_container_width=True)
            st.caption("Lower WACC generally increases value; higher terminal growth generally increases value.")

            st.markdown("#### Sensitivity analysis — WACC / exit EV/EBITDA multiple")
            multiples = np.linspace(max(1.0, exit_multiple - 2.0), exit_multiple + 2.0, 5)
            exit_sens = sensitivity_exit(
                base_revenue, growths, ebit_margin, tax_rate, da_rev, capex_rev,
                nwc_rev, net_debt, diluted_shares, wacc_values, multiples, terminal_growth
            )
            st.dataframe(exit_sens.style.format("${:,.2f}", na_rep="N/M"), use_container_width=True)

            st.markdown("#### Valuation football field")
            pg_vals = pg_sens.to_numpy(dtype=float)
            ex_vals = exit_sens.to_numpy(dtype=float)
            ranges = [
                ("Perpetuity Growth DCF", np.nanmin(pg_vals), np.nanmax(pg_vals), dcf_out["PG Implied Share Price"]),
                ("Exit Multiple DCF", np.nanmin(ex_vals), np.nanmax(ex_vals), dcf_out["Exit Implied Share Price"]),
            ]
            fig, ax = plt.subplots(figsize=(9, 3.4))
            for y, (label, lo, hi, mid) in enumerate(ranges):
                ax.hlines(y, lo, hi, linewidth=8)
                ax.plot(mid, y, "o")
            if pd.notna(px):
                ax.axvline(px, linestyle="--", linewidth=1.5)
            ax.set_yticks(range(len(ranges)))
            ax.set_yticklabels([r[0] for r in ranges])
            ax.set_xlabel("Implied value per share ($)")
            ax.set_title("DCF Valuation Football Field")
            ax.grid(axis="x", alpha=0.25)
            st.pyplot(fig, clear_figure=True)

            st.info(
                "Historical ratios inform assumptions; they do not mechanically set them. "
                "Students must explain why forecast growth, margins, capex, NWC, and discount-rate assumptions "
                "differ from historical experience."
            )

            dcf_inputs = {
                "base_revenue_mm": base_revenue,
                "growths": growths,
                "ebit_margin": ebit_margin,
                "tax_rate": tax_rate,
                "da_revenue": da_rev,
                "capex_revenue": capex_rev,
                "change_nwc_revenue": nwc_rev,
                "wacc": wacc,
                "terminal_growth": terminal_growth,
                "exit_ev_ebitda_multiple": exit_multiple,
                "net_debt_mm": net_debt,
                "diluted_shares_mm": diluted_shares,
            }
            wacc_inputs = {
                "rf": rf, "beta": beta, "erp": erp, "cost_debt": pretax_cod,
                "tax": wacc_tax, "market_equity": market_equity, "debt": debt_for_wacc,
                "calculated_wacc": wacc_calc,
            }

            # Store for later tabs in same rerun.
            st.session_state[f"dcf_inputs_{ticker}"] = dcf_inputs
            st.session_state[f"wacc_inputs_{ticker}"] = wacc_inputs
            st.session_state[f"dcf_outputs_{ticker}"] = dcf_out


# -----------------------------
# Shared source evidence rows
# -----------------------------
source_rows = [
    {"Source": "SEC Company Tickers", "URL": SEC_TICKERS_URL, "Section/Page": "", "Evidence/Note": "Ticker ↔ CIK mapping"},
    {"Source": "SEC Submissions", "URL": submissions_url, "Section/Page": "", "Evidence/Note": "Latest filing metadata"},
    {"Source": "SEC Companyfacts/XBRL", "URL": facts_url, "Section/Page": "", "Evidence/Note": "Historical accounting facts"},
]
if filing_url:
    source_rows.append({"Source": "Latest 10-K", "URL": filing_url, "Section/Page": "", "Evidence/Note": "Reconciliation, revenue drivers, Item 1A risks"})
source_rows += [
    {"Source": "U.S. Treasury Daily Rates", "URL": TREASURY_URL.format(year=date.today().year), "Section/Page": "10 Yr", "Evidence/Note": "Risk-free rate"},
    {"Source": "Damodaran ERP", "URL": DAMODARAN_ERP_URL, "Section/Page": "Current implied ERP", "Evidence/Note": "Equity risk premium"},
    {"Source": "yfinance / Yahoo market data", "URL": f"https://finance.yahoo.com/quote/{ticker}/", "Section/Page": "", "Evidence/Note": "Market price / market-cap reference; beta cross-check"},
]


# -----------------------------
# Tab 7 Excel
# -----------------------------
with tabs[6]:
    st.subheader("Module 7 — Excel Model Generator")
    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete.")
    else:
        dcf_inputs_ss = st.session_state.get(f"dcf_inputs_{ticker}")
        wacc_inputs_ss = st.session_state.get(f"wacc_inputs_{ticker}")
        dcf_outputs_ss = st.session_state.get(f"dcf_outputs_{ticker}")
        if not dcf_inputs_ss or not wacc_inputs_ss or not dcf_outputs_ss:
            st.warning("Open the DCF-WACC tab once and review the assumptions before generating Excel.")
        else:
            # Re-key for workbook helper.
            dcf_excel = {
                "base_revenue": dcf_inputs_ss["base_revenue_mm"],
                "growths": dcf_inputs_ss["growths"],
                "ebit_margin": dcf_inputs_ss["ebit_margin"],
                "tax_rate": dcf_inputs_ss["tax_rate"],
                "da_rev": dcf_inputs_ss["da_revenue"],
                "capex_rev": dcf_inputs_ss["capex_revenue"],
                "nwc_rev": dcf_inputs_ss["change_nwc_revenue"],
                "terminal_growth": dcf_inputs_ss["terminal_growth"],
                "exit_multiple": dcf_inputs_ss["exit_ev_ebitda_multiple"],
                "net_debt": dcf_inputs_ss["net_debt_mm"],
                "shares": dcf_inputs_ss["diluted_shares_mm"],
            }
            xlsx = build_excel(
                ticker, company_name, fin, ratios, mapping, dcf_excel, wacc_inputs_ss,
                source_rows, px
            )
            st.download_button(
                "Download Excel model",
                data=xlsx,
                file_name=f"{ticker}_equity_research_model.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.markdown(
                "**Formula-control test:** open the workbook → DCF sheet → change a yellow Revenue Growth, "
                "EBIT Margin, Terminal Growth, or WACC input (WACC inputs are on the WACC sheet) → confirm "
                "the implied share-price formulas recalculate."
            )
            st.write(
                "Workbook sheets: Historical, Ratios, WACC, DCF, XBRL Mapping, Sources. "
                "Green cells are formulas; yellow cells are editable assumptions."
            )

            evidence_df = pd.DataFrame(source_rows)
            st.download_button(
                "Download source evidence log template",
                data=evidence_df.to_csv(index=False).encode("utf-8"),
                file_name=f"{ticker}_source_evidence_log.csv",
                mime="text/csv",
            )
            st.caption(
                "Students should add filing page/section references and screenshot filenames manually. "
                "The SEC HTML filing is the primary source; screenshots are evidence artifacts, not substitutes for source verification."
            )


# -----------------------------
# Tab 8 Memo
# -----------------------------
with tabs[7]:
    st.subheader("Module 8 — AI Memo Agent")
    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete.")
    else:
        dcf_inputs_ss = st.session_state.get(f"dcf_inputs_{ticker}")
        dcf_outputs_ss = st.session_state.get(f"dcf_outputs_{ticker}")
        if not dcf_inputs_ss or not dcf_outputs_ss:
            st.warning("Complete/review the DCF tab first.")
        else:
            item1 = extract_item_section(filing_text, "1", "1A")
            item7 = extract_item_section(filing_text, "7", "7A")
            rev_source_text = (item1 + " " + item7).strip() or filing_text
            revenue_evidence = evidence_snippets(
                rev_source_text,
                ["price", "pricing", "volume", "shipment", "mix", "demand", "capacity", "sales", "revenue", "tons"],
                max_snippets=10,
            )
            item1a = extract_item_section(filing_text, "1A", "1B") or extract_item_section(filing_text, "1A", "2")
            risk_evidence = evidence_snippets(
                item1a or filing_text,
                ["risk", "commodity", "price", "demand", "cyber", "environment", "regulation",
                 "competition", "supply", "energy", "labor", "capital", "debt", "interest"],
                max_snippets=8,
            )
            company_info = {
                "company": company_name, "ticker": ticker, "CIK": str(cik).zfill(10),
                "market_price_reference": px, "sector": market.get("sector"),
                "industry": market.get("industry"),
            }
            prompt = build_memo_prompt(
                company_name, ticker, company_info, fin, ratios, dcf_inputs_ss, dcf_outputs_ss,
                revenue_evidence, risk_evidence, filing_url, source_rows
            )
            st.write(
                "No paid API is required. Copy this prompt into your institution-approved AI tool. "
                "The app supplies evidence and calculations; the AI supplies synthesis only."
            )
            st.text_area("Equity research memo prompt", prompt, height=600)
            st.info("Target student deliverable: a 4–6 page first-draft equity research note, then independently audited.")


# -----------------------------
# Tab 9 Audit
# -----------------------------
with tabs[8]:
    st.subheader("Module 9 — Hallucination Audit Agent")
    if not gate_passed():
        st.error("Accounting-data reconciliation is incomplete.")
    else:
        st.markdown("**Workflow:** AI Claim → Classify → Find Source → Verify → Correct → Document")
        memo = st.text_area(
            "Paste the AI-generated memo here",
            key=f"memo_{ticker}",
            height=320,
            placeholder="Paste the junior-analyst AI first draft...",
        )
        audit_df = claim_candidates(memo, limit=20)
        if not audit_df.empty:
            st.markdown("#### 15–20 material-claim audit table")
            edited = st.data_editor(
                audit_df,
                use_container_width=True,
                num_rows="dynamic",
                column_config={
                    "Claim Type": st.column_config.SelectboxColumn(
                        "Claim Type",
                        options=[
                            "Historical Fact", "Calculation", "Management Statement",
                            "Analytical Interpretation", "Forecast Assumption", "Valuation Output"
                        ],
                    ),
                    "Verification Status": st.column_config.SelectboxColumn(
                        "Verification Status",
                        options=[
                            "VERIFIED", "PARTIALLY SUPPORTED / QUALIFY",
                            "UNSUPPORTED", "CONTRADICTED"
                        ],
                    ),
                },
                key=f"audit_editor_{ticker}",
            )
            st.download_button(
                "Download AI audit appendix CSV",
                data=edited.to_csv(index=False).encode("utf-8"),
                file_name=f"{ticker}_AI_audit_appendix.csv",
                mime="text/csv",
            )

            causal_terms = [
                "because", "due to", "driven by", "primarily", "reflecting", "resulted from",
                "management expects", "management believes", "management indicated",
                "the company expects", "guidance suggests", "management anticipates",
            ]
            flagged = [t for t in causal_terms if re.search(re.escape(t), memo, re.I)]
            if flagged:
                st.warning("Causal/management-attribution language detected: " + ", ".join(flagged))

            st.markdown("#### Forecast audit rule")
            st.write(
                'If AI says "Revenue is expected to grow 5% next year," ask: **Expected by whom?** '
                'If 5% came from the student model, rewrite: '
                '"Our base-case model assumes 5% revenue growth next year."'
            )
            st.markdown("#### Second-AI preliminary audit prompt")
            st.text_area("Optional pre-audit prompt", second_audit_prompt(memo), height=360)
        else:
            st.info("Paste a memo to generate candidate claims for manual classification and verification.")

        st.markdown("#### Detailed AI error log — required when problems are present")
        error_template = pd.DataFrame([
            {"Original AI statement": "", "Evidence checked": "", "Problem": "",
             "Verification classification": "", "Corrected statement": "", "Likely reason AI made the mistake": ""},
            {"Original AI statement": "", "Evidence checked": "", "Problem": "",
             "Verification classification": "", "Corrected statement": "", "Likely reason AI made the mistake": ""},
            {"Original AI statement": "", "Evidence checked": "", "Problem": "",
             "Verification classification": "", "Corrected statement": "", "Likely reason AI made the mistake": ""},
        ])
        st.data_editor(error_template, use_container_width=True, num_rows="dynamic", key=f"errors_{ticker}")

        st.markdown("#### Source provenance rule")
        st.write(
            "Every material statement in the final note should answer at least one: "
            "(1) Where is the source? (2) Where is the calculation? "
            "(3) Where is the model assumption? (4) Why is this my analytical judgment? "
            "If none applies, remove or rewrite it."
        )
        st.error(
            "AI may draft and may pre-audit. The student must manually verify and correct. "
            "Do not submit a memo merely because an AI says its own output is correct."
        )

st.divider()
st.caption(
    "Educational use only; not investment advice. SEC filing data can require company-specific XBRL mapping. "
    "Market-data convenience sources can be incomplete. Always reconcile primary-source accounting data."
)
